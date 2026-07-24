"""Derive the hf_bill seed from Household-Monthly-Nut-v2.xlsx (Sprint 3R item a).

The Register carries name/tier/amount/decision, but NOT the two fields PRD 6.2
requires for matching: which account pays, and when. Both are recoverable — the
Raw Checking / Raw Apple Card sheets tag every transaction with its 'Assigned
group', i.e. the Register line it rolls up to. So we derive:

  payment account   <- which raw sheet the group's transactions came from
  due day-of-month  <- median posting day across observed cycles
  match pattern     <- longest common prefix of the dominant merchant string
  expected amount   <- median observed amount (Register 'Est monthly' kept alongside)

Output is a REVIEW CSV, not a direct write. John verifies/edits it, then
seed_bills.py loads the reviewed file. Derivation never silently becomes truth.

Usage:
  python derive_bill_seed.py "..\\..\\Household-Monthly-Nut-v2.xlsx" bill_seed_review.csv
"""
import csv
import datetime as dt
import re
import statistics
import sys
from collections import Counter, defaultdict

import openpyxl

WB_PATH = sys.argv[1]
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else "bill_seed_review.csv"

# Register rows that are real obligations (a biller + a cadence) rather than
# spend rollups. Everything else in tiers 2/3 is kind=category: it lands in the
# table so the monthly nut stays queryable in one place, but the matcher skips
# it -- a category has no due date, so it could only ever produce false MISSED.
BILL_OVERRIDES = {
    "Water delivery (Primo)",
    "Milk delivery (Alpenrose/Smith Bros)",
    "Home Chef (CANCELLED)",
}

# Bills whose Register line aggregates several distinct subscriptions. They seed
# as estimate-only (hf_matchmode=none) because no single merchant+amount pair can
# match them; splitting them into real bills is a John decision, not a derivation.
AGGREGATES = {
    "Apple services (checking side)",
    "Apple services (card side)",
    "Xbox/Microsoft (checking side)",
    "Xbox/Microsoft (card side)",
    "Annual subs: Plex/1Password/MyQ (card)",
    "24 Hour Fitness (annual + day fees)",
    "USAA Visa payment (proxy)",  # Visa txns absent from sync; payments-only proxy
}

# Register name -> raw-sheet 'Assigned group' where the two drifted apart.
# Verified exhaustively: these are the ONLY mismatches. Every other unmatched raw
# group is income (payroll, VA, interest) or an excluded card-payment transfer,
# neither of which belongs in the bill registry.
ALIASES = {"Seed probiotics (CANCELLED)": "Seed probiotics (checking + card)"}

# Apple Card statement latency: a card-side bill cannot be confirmed until the
# monthly CSV is exported. Missed-detection waits this long before firing.
LATENCY_DAYS = {"checking": 3, "applecard": 35, "mixed": 35, "unknown": 35}

CADENCE_HINTS = [
    (re.compile(r"every 2 months", re.I), "bimonthly"),
    (re.compile(r"every 3 months|quarterly", re.I), "quarterly"),
    # Deliberately NOT matching bare 'renew': Tesla FSD's note ("was set to renew
    # 7/27") is a monthly sub and was mis-derived as annual on the first pass.
    (re.compile(r"/yr|\bannual", re.I), "annual"),
]

# Facts John supplied that the transaction window is too short to reveal. Dollar
# Shave bills quarterly (only one charge fell inside the April-July data, so the
# derivation could not see the 3-month gap); CrunchLabs was cancelled after June.
# Encoded here so a regenerate preserves them rather than reverting to a guess.
FREQUENCY_OVERRIDES = {"Dollar Shave Club": "quarterly"}
STATUS_OVERRIDES = {"CrunchLabs": "cancelled"}

# Share of a bill's own transactions the pattern must catch before we trust it to
# find every cycle. Below this the pattern is precise but leaky -- some cycles
# would go unmatched and surface as false MISSED.
MIN_COVERAGE = 0.6

# Cap on distinct descriptors per bill. Beyond a couple of aliases the line is
# not one biller with a renamed descriptor, it is a spend category in disguise.
MAX_PATTERNS = 3


def slug(name):
    # 92, not 100: hf_billinstance.hf_instancekey is "<billkey>|YYYY-MM" and must
    # fit Text(100). Longest real key today is 35 chars, but the alternate key
    # would reject silently-truncated collisions, so leave the room.
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")[:92]


def tier_of(raw):
    if not raw:
        return ""
    return raw.split(" - ")[0].strip()


def read_register(wb):
    ws = wb["Register"]
    rows = []
    for r in ws.iter_rows(min_row=7, values_only=True):
        name = r[0]
        if not name or str(name).startswith("TOTAL"):
            continue
        rows.append({
            "name": str(name),
            "tier_raw": str(r[1] or ""),
            "amount_type": str(r[2] or ""),
            "est_monthly": r[3],
            "decision": str(r[8] or ""),
            "notes": str(r[10] or "").replace("None", ""),
        })
    return rows


def as_date(v):
    """Raw sheets store dates as ISO text in some rows, real datetimes in others."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def read_raw(wb):
    """group -> list of (date, merchant, amount_abs, account)"""
    obs = defaultdict(list)
    for sheet, account, amt_col, grp_col in (
        ("Raw Checking", "checking", 3, 4),
        ("Raw Apple Card", "applecard", 4, 5),
    ):
        for r in wb[sheet].iter_rows(min_row=2, values_only=True):
            group, amount, date = r[grp_col], r[amt_col], as_date(r[0])
            if not group or amount is None or date is None:
                continue
            obs[str(group)].append((date, str(r[1] or ""), abs(float(amount)), account))
    return obs


# Precision is measured against observed history, so a very short prefix can test
# clean today and over-match later -- "Monthl" and "Google" both scored 1.0 purely
# because no counterexample had posted yet. Keeping prefixes long enough to stay
# distinctive costs nothing now that a bill can carry several of them.
MIN_PREFIX = 10


def shortest_precise_prefix(merchant, group, obs):
    """Shortest prefix of `merchant` that still catches only `group`.

    Descriptors carry per-transaction noise (order ids, store numbers), so an
    exact match would miss every cycle. Shortening until precision would break
    gives the widest prefix that is still unambiguous -- "Patreon" rather than
    "Patreon* Membership Internet", which is what lets it also catch the
    "Patreon Patreon* Membershwww.patreon.cCA" variant of the same charge.
    """
    best = None
    for n in range(MIN_PREFIX, len(merchant) + 1):
        if precision(merchant[:n], group, obs) == 1.0:
            best = merchant[:n]
            break
    return (best or merchant).strip()[:80]


def merchant_patterns(merchants, group, obs):
    """A set of prefixes covering the bill, joined with '|'.

    One pattern is not always enough. The mortgage posts as 'NSM DBAMR.COOPER MR
    COOPER' before June 2026 and 'M & T MORTGAGE MTG PYT' after -- the servicer
    changed, and no single prefix spans both. Treating the pattern as a set means
    a renamed biller is a second pattern, not a run of false MISSED.
    """
    if not merchants:
        return "", 0.0
    prefixes = []
    for merchant, _ in Counter(merchants).most_common():
        p = shortest_precise_prefix(merchant, group, obs)
        # Drop any prefix already covered by a shorter one we kept.
        if not any(p.startswith(k) for k in prefixes):
            prefixes = [k for k in prefixes if not k.startswith(p)] + [p]
        if len(prefixes) >= MAX_PATTERNS:
            break
    covered = sum(1 for m in merchants if any(m.startswith(p) for p in prefixes))
    return "|".join(prefixes), round(covered / len(merchants), 2)


def precision(pattern, group, obs):
    """Fraction of the whole raw corpus matching `pattern` that is really `group`.

    Pattern length is a poor proxy for ambiguity -- "Hulu" is 4 characters and
    perfectly unambiguous, while "CITY OF" is 7 and not. The corpus answers the
    question directly: if this pattern also catches another bill's transactions,
    merchant alone cannot bind a match.
    """
    if not pattern:
        return 0.0
    p = pattern.lower()
    hit = own = 0
    for g, rows in obs.items():
        for _, merchant, _, _ in rows:
            if merchant.lower().startswith(p):
                hit += 1
                own += (g == group)
    return round(own / hit, 2) if hit else 0.0


def cadence(name, notes, dates):
    if name in FREQUENCY_OVERRIDES:
        return FREQUENCY_OVERRIDES[name]
    for rx, val in CADENCE_HINTS:
        if rx.search(notes):
            return val
    months = {(d.year, d.month) for d in dates}
    if len(months) >= 2 and len(dates) / len(months) >= 1.8:
        return "monthly"  # multiple hits per month -> still monthly, amount varies
    return "monthly" if months else "unknown"


def main():
    wb = openpyxl.load_workbook(WB_PATH, data_only=True)
    register, obs = read_register(wb), read_raw(wb)

    out = []
    for row in register:
        name = row["name"]
        tier = tier_of(row["tier_raw"])
        hits = obs.get(ALIASES.get(name, name), [])
        dates = [d for d, _, _, _ in hits]
        amounts = [a for _, _, a, _ in hits]
        accounts = {acct for _, _, _, acct in hits}
        merchants = [m for _, m, _, _ in hits]

        if tier == "One-off (excluded)":
            kind = "excluded"
        elif tier == "1" or name in BILL_OVERRIDES or (
            tier == "3" and row["amount_type"] == "Fixed"
        ):
            kind = "bill"
        else:
            kind = "category"

        account = ("mixed" if len(accounts) > 1
                   else next(iter(accounts)) if accounts else "unknown")
        group_key = ALIASES.get(name, name)
        pattern, share = merchant_patterns(merchants, group_key, obs)
        freq = cadence(name, row["notes"], dates)
        days = sorted(d.day for d in dates)
        cancelled = (row["decision"] == "Cut" or "CANCELLED" in name.upper()
                     or STATUS_OVERRIDES.get(name) == "cancelled")

        prec = min([precision(p, group_key, obs) for p in pattern.split("|")] or [0.0])
        if kind != "bill" or name in AGGREGATES:
            match_mode = "none"
        elif prec < 1.0:
            # Pattern also catches another bill; amount has to disambiguate.
            match_mode = "merchant+amount"
        elif share >= MIN_COVERAGE:
            match_mode = "merchant"
        else:
            match_mode = "review"

        out.append({
            "bill_key": slug(name),
            "name": name,
            "kind": kind,
            "tier": tier,
            "status": "cancelled" if cancelled else "active",
            "amount_type": row["amount_type"].lower(),
            # Two different numbers, and conflating them was the first pass's bug.
            # The Register's 'Est monthly' is a monthly-equivalent for the nut:
            # garbage shows 47.36 because 94.72 falls every second month. Matching
            # needs the per-cycle charge, so a non-monthly bill takes the observed
            # median instead -- otherwise every garbage cycle reads as +100% drift.
            "expected_amount": (
                row["est_monthly"] if freq == "monthly" and row["est_monthly"]
                else round(statistics.median(amounts), 2) if amounts
                else row["est_monthly"] or ""),
            "monthly_equivalent": row["est_monthly"] if row["est_monthly"] is not None else "",
            "observed_median": round(statistics.median(amounts), 2) if amounts else "",
            "frequency": freq,
            "due_day": int(statistics.median(days)) if days else "",
            "due_day_spread": f"{days[0]}-{days[-1]}" if days else "",
            # Cadence anchor. Bimonthly and annual bills need a phase, not just a
            # day -- garbage is every 2 months from *some* month, GoDaddy renews
            # in *some* month. Stepping the frequency from the first observed
            # charge gives monthly, bimonthly and annual one uniform rule.
            "anchor_date": min(dates).isoformat() if dates else "",
            "payment_account": account,
            "latency_days": LATENCY_DAYS[account],
            "match_mode": match_mode,
            "match_pattern": pattern,
            "pattern_coverage": share,
            "pattern_precision": prec,
            "observations": len(hits),
            "variance_tolerance_pct": 15,
            "end_date": "",
            "notes": row["notes"],
        })

    with open(OUT_PATH, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(out[0].keys()))
        w.writeheader()
        w.writerows(out)

    kinds = Counter(r["kind"] for r in out)
    modes = Counter(r["match_mode"] for r in out if r["kind"] == "bill")
    print(f"wrote {OUT_PATH}: {len(out)} rows  {dict(kinds)}")
    print(f"  bill match_mode: {dict(modes)}")
    print(f"  no observations: {[r['name'] for r in out if not r['observations']]}")


if __name__ == "__main__":
    main()
